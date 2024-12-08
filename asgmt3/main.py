import os
import logging
import argparse
import openpyxl
import docx2pdf
import PyPDF2

logging.basicConfig(
    level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s"
)


def process_excel_files(input_folder: str, output_folder: str) -> None:
    """
    Process Excel files in the given input folder and create a summary Excel file in the output folder.

    This function reads all Excel files (.xlsx) in the input folder, processes each sheet as per the predefined configurations,
    and creates a summary Excel file with aggregated data and charts in the output folder.

    Args:
        input_folder (str): The folder containing the Excel files to be processed.
        output_folder (str): The folder where the summary Excel file will be saved.

    Returns:
        None
    """

    # Define the configuration for each sheet
    sheet_configurations = {
        "论文": {
            "default_columns": [
                "论文名称",
                "作者列表",
                "作者单位",
                "发表日期",
                "论文级别",
            ],
            "unique_key_column": "论文名称",
            "date_column_name": "发表日期",
            "column_widths": [
                50,
                20,
                50,
                20,
                20,
            ],
        },
        "专利": {
            "default_columns": [
                "专利名称",
                "专利授权号",
                "被授权人",
                "被授权人单位",
                "授权日期",
            ],
            "unique_key_column": "专利名称",
            "date_column_name": "授权日期",
            "column_widths": [
                50,
                20,
                20,
                20,
                20,
            ],
        },
    }

    # Extract sheet names from the configuration
    sheet_names = list(sheet_configurations.keys())

    # Initialize a summary dictionary for storing data
    data_summary: dict[str, dict[str, list[str]]] = {
        sheet_name: {} for sheet_name in sheet_names
    }

    # List all .xlsx files in the input folder
    xlsx_files = [f for f in os.listdir(input_folder) if f.endswith(".xlsx")]

    # Check if there are any Excel files in the input folder
    if not xlsx_files:
        logging.error("No Excel files found in the input folder.")
        return

    # Process each Excel file
    for file in xlsx_files:
        file_path = os.path.join(input_folder, file)
        logging.info(f"Processing file: {file_path}")
        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True)
            # Process each sheet in the workbook
            for sheet_name in sheet_names:
                worksheet = workbook[sheet_name]
                first_row = next(worksheet.iter_rows(values_only=True))
                # Check if the first row matches the unique key column
                if (
                    first_row[0]
                    != sheet_configurations[sheet_name]["unique_key_column"]
                ):
                    logging.warning(
                        f"No header found in {file} sheet {sheet_name}. Using default columns."
                    )
                    data_summary[sheet_name][first_row[0]] = first_row[1:]
                # Iterate over the rows in the sheet
                for row in worksheet.iter_rows(min_row=2, values_only=True):
                    data_summary[sheet_name][row[0]] = row[1:]
        except Exception as e:
            logging.error(f"Error processing {file}: {e}")

    # Create a new workbook for the summary
    summary_workbook = openpyxl.Workbook()
    summary_workbook.remove(summary_workbook["Sheet"])
    # Populate the summary workbook
    for sheet_name, data in data_summary.items():
        summary_worksheet = summary_workbook.create_sheet(title=sheet_name + "汇总表")
        summary_worksheet.append(sheet_configurations[sheet_name]["default_columns"])

        # Set column widths
        for idx, column_width in enumerate(
            sheet_configurations[sheet_name]["column_widths"], start=1
        ):
            summary_worksheet.column_dimensions[
                openpyxl.utils.get_column_letter(idx)
            ].width = column_width

        # Append data to the summary worksheet
        for unique_key, row_data in data.items():
            summary_worksheet.append([unique_key] + list(row_data))

        # Count occurrences of each year
        date_column_index = sheet_configurations[sheet_name]["default_columns"].index(
            sheet_configurations[sheet_name]["date_column_name"]
        )
        year_counts = {}
        for unique_key, row_data in data.items():
            year = str(row_data[date_column_index - 1].year)
            year_counts[year] = year_counts.get(year, 0) + 1

        # Append year counts to the summary worksheet
        summary_worksheet.append(["年份", "数量"])
        for year, count in year_counts.items():
            summary_worksheet.append([year, count])

        # Create a chart for the year counts
        stat_data = openpyxl.chart.Reference(
            summary_worksheet,
            min_col=2,
            min_row=len(data) + 3,
            max_row=len(data) + 2 + len(year_counts),
        )

        categories = openpyxl.chart.Reference(
            summary_worksheet,
            min_col=1,
            min_row=len(data) + 3,
            max_row=len(data) + 2 + len(year_counts),
        )

        series = openpyxl.chart.Series(stat_data, title="年度数量统计")

        chart = openpyxl.chart.BarChart()
        chart.title = "年度数量统计"
        chart.x_axis.title = "年份"
        chart.y_axis.title = "数量"

        chart.append(series)
        chart.set_categories(categories)
        summary_worksheet.add_chart(chart, "A20")

    # Save the summary workbook to the output folder with exception handling
    try:
        output_file_path = os.path.join(output_folder, "汇总Excel.xlsx")
        summary_workbook.save(output_file_path)
        logging.info(f"Summary workbook saved successfully to {output_file_path}")
    except PermissionError:
        logging.error(
            f"Permission denied: Unable to save the summary workbook to {output_file_path}"
        )
    except Exception as e:
        logging.error(f"An error occurred while saving the summary workbook: {e}")


def convert_word_files_to_pdfs(input_folder) -> None:
    """
    Convert all Word files (.docx) in the given input folder to PDF files in the same folder.

    This function uses the docx2pdf library to convert the Word files to PDF files.
    The PDF files are saved in the same folder as the input Word files, with the same name but with a .pdf extension instead of .docx.

    :param input_folder: The folder containing the Word files to be converted
    :return: None
    """

    logging.info("Converting Word files to PDF...")
    docx_files = [f for f in os.listdir(input_folder) if f.endswith(".docx")]

    # Iterate over the list of Word files
    for file in docx_files:
        docx_path = os.path.join(input_folder, file)
        pdf_path = os.path.join(input_folder, file.replace(".docx", ".pdf"))

        try:
            # Convert the Word file to a PDF file
            docx2pdf.convert(docx_path, pdf_path)
            logging.info(f"Converted {docx_path} to {pdf_path}")
        except Exception as e:
            # Log any errors that occur during the conversion process
            logging.error(f"Error converting {docx_path}: {e}")


def merge_pdfs_with_watermark(input_folder, output_folder, watermark_path) -> None:
    """
    Merge PDF files in the given input folder with a watermark PDF file, and save the merged PDF files to the given output folder.

    This function reads all PDF files in the input folder, merges them with the given watermark PDF file, and saves the merged PDF files to the output folder.
    The merged PDF files are saved with the same name as the original PDF files, but with "_watermark" appended to the file name.

    :param input_folder: The folder containing the PDF files to be merged
    :param output_folder: The folder where the merged PDF files will be saved
    :param watermark_path: The path to the watermark PDF file
    :return: None
    """

    pdf_writer = PyPDF2.PdfWriter()

    try:
        # Get the list of PDF files in the input folder
        pdf_files = [f for f in os.listdir(input_folder) if f.endswith(".pdf")]

        # Open the watermark PDF file
        with open(watermark_path, "rb") as watermark_file:
            watermark_reader = PyPDF2.PdfReader(watermark_file)
            watermark_page = watermark_reader.pages[0]

            # Iterate over the list of PDF files
            for file in pdf_files:
                if file != os.path.basename(watermark_path):
                    pdf_path = os.path.join(input_folder, file)

                    try:
                        # Open the PDF file and read its pages
                        with open(pdf_path, "rb") as pdf_file:
                            pdf_reader = PyPDF2.PdfReader(pdf_file)
                            for page_num in range(len(pdf_reader.pages)):
                                page = pdf_reader.pages[page_num]
                                # Merge the watermark page with the current page
                                page.merge_page(watermark_page)
                                # Add the merged page to the output PDF file
                                pdf_writer.add_page(page)
                    except Exception as e:
                        # Log any errors that occur during the merging process
                        logging.error(f"Error merging {pdf_path}: {e}")
    except Exception as e:
        # Log any errors that occur during the merging process
        logging.error(f"Error merging PDFs with watermark: {e}")

    try:
        # Save the merged PDF file to the output folder
        output_path = os.path.join(output_folder, "打印文档汇总.pdf")
        with open(output_path, "wb") as out:
            pdf_writer.write(out)
        logging.info(f"Created merged PDF with watermark at {output_path}")
    except Exception as e:
        # Log any errors that occur during the merging process
        logging.error(f"Error creating merged PDF: {e}")


if __name__ == "__main__":
    # Parse command line arguments
    parser = argparse.ArgumentParser(
        description="处理 Excel 文件或转换 Word 文档为 PDF"
    )
    parser.add_argument(
        "--no_process_excel",
        action="store_true",
        help="禁用汇总 Excel 文件功能",
    )
    parser.add_argument(
        "--no_convert_word",
        action="store_true",
        help="禁用转换 Word 文件为 PDF 并合并 PDF 文件功能",
    )

    # Parse arguments
    args: argparse.Namespace = parser.parse_args()

    # Set input and output folders
    input_folder: str = "files"
    output_folder: str = os.path.join("files", "汇总")
    watermark_path: str = os.path.join("files", "水印文件.pdf")

    # Create output folder if it does not exist
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    # Indicator to track if any processing is performed
    processing_done = False

    # Process Excel files if not disabled
    if not args.no_process_excel:
        process_excel_files(input_folder, output_folder)
        processing_done = True

    # Convert Word files to PDF and merge PDF files with watermark if not disabled
    if not args.no_convert_word:
        convert_word_files_to_pdfs(input_folder)
        merge_pdfs_with_watermark(input_folder, output_folder, watermark_path)
        processing_done = True

    # Print help if no processing is done
    if not processing_done:
        parser.print_help()
